"""Importa lotes históricos del Excel "LOTES RECIBIDOS Y DETERMINACIÓN DE SULFITOS".

Uso:
    python -m scripts.import_excel <ruta_xlsx>
    python -m scripts.import_excel <ruta_xlsx> --dry-run
    python -m scripts.import_excel <ruta_xlsx> --limit 50

El script:
  - Lee las 4 hojas (ENERO, FEBRERO, MARZO, ABRIL).
  - Crea catálogos nuevos (proveedores, procedencias, PSCs, logísticas,
    tratadores) sobre la marcha — find-or-create.
  - Idempotente: si un lote (lot_code + lot_year) ya existe, lo salta.
  - Crea por cada fila: lot, reception + reception_lot, quality_analysis,
    analysis_lot, analysis_color (cocido), analysis_flavor (cocido) y
    UNA analysis_sampling (sampling_index=1) con todos los defectos como
    sampling_defects (porque el Excel no separa los 3 muestreos).
  - Status del análisis se importa como 'borrador'. Sin decisión.
"""
import argparse
import logging
import re
import sys
import unicodedata
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

# Silenciar logs ruidosos de SQLAlchemy en el script
logging.getLogger("sqlalchemy.engine").setLevel(logging.WARNING)

from app.core.db import SessionLocal, engine

# Silenciar SQL echo (queremos solo nuestros prints, no la query log)
engine.echo = False
from app.models.analyses import (
    AnalysisColor,
    AnalysisFlavor,
    AnalysisLot,
    AnalysisSampling,
    QualityAnalysis,
    SamplingDefect,
)
from app.models.attachments import Attachment  # noqa: F401 — registra modelo
from app.models.auth import User
from app.models.catalogs import (
    Chemical,
    Color,
    Defect,
    Flavor,
    Intensity,
    LogisticsCompany,
    LotCategory,
    Origin,
    Plant,
    Pond,
    Supplier,
    Treater,
)
from app.models.operations import Lot, Reception, ReceptionLot, lot_treaters_table


# Cabecera está en la fila 4 del Excel (0-indexed = 3).
# Las filas 0-3 son títulos vacíos; los datos empiezan en la 4.
HEADER_ROW_IDX = 3

# Mapa de columnas (0-indexed) según la cabecera del Excel
COLS = {
    "turno": 0,
    "planta": 1,
    "hora_analisis": 2,
    "fecha_recepcion": 3,
    "lote": 4,
    "proveedor": 5,
    "procedencia": 6,
    "psc": 7,
    "tipo_producto": 8,
    "lbs": 9,
    "gr_cc": 10,
    "c_kg": 11,
    "gr_sc": 12,
    "c_kg2": 13,
    "color": 14,
    "sabor": 15,
    "tratador": 16,
    "metabisulfito": 17,
    "so2_global": 18,
    "logistica": 19,
    "pct_defectos": 20,
    "observacion": 35,
}

# Excel column index → defect_code en catálogo `defects`
DEFECT_COL_MAP = {
    21: "cabeza_roja",
    22: "cabeza_naranja",
    23: "cabeza_floja",
    24: "cabeza_descolgada",
    25: "cabeza_reventada",
    26: "flacido",
    27: "mudado",
    28: "picado",
    29: "atq_bacteriano",
    30: "quebrado",
    31: "semi_rosado",
    32: "melanosis",
    33: "deforme",
    34: "deshidratado",
}


# ─────────────────────────────────────────────────────────────────────
# Parsing helpers
# ─────────────────────────────────────────────────────────────────────

def normalize(s):
    """Strip + uppercase + colapsa espacios + quita acentos. None devuelve None."""
    if s is None:
        return None
    s = str(s).strip().upper()
    # Quita acentos: "CARACTERÍSTICO" → "CARACTERISTICO", "DURÁN" → "DURAN"
    s = "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )
    n = " ".join(s.split())
    return n if n else None


def parse_decimal(v):
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except Exception:
        return None


def parse_date_cell(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                pass
    return None


def parse_time_cell(v):
    if v is None:
        return None
    if isinstance(v, time):
        return v
    if isinstance(v, datetime):
        return v.time()
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(s, fmt).time()
            except ValueError:
                pass
    return None


def parse_treaters(text):
    """'FREIRE Y LA TORRES' → ['FREIRE', 'LA TORRES']; 'PROVEEDOR' → ['PROVEEDOR']."""
    if not text:
        return []
    s = str(text).strip()
    # Acepta separadores: " Y ", " y ", " / ", "/"
    parts = re.split(r"\s+Y\s+|\s+y\s+|\s*/\s*", s)
    return [normalize(p) for p in parts if p and p.strip()]


def parse_so2_values(cell):
    """Devuelve lista de Decimal (1-3 valores) desde la celda de SO2 GLOBAL.

    Casos:
      80                                  → [80]
      "80"                                → [80]
      "123,05    85,60    65,20"          → [123.05, 85.60, 65.20]
      "93,77   50,60   71,20"             → [93.77, 50.60, 71.20]
    Si hay más de 3 valores, se truncan a los primeros 3.
    Convierte coma decimal española a punto.
    """
    if cell is None:
        return []
    if isinstance(cell, (int, float, Decimal)):
        return [Decimal(str(cell))]
    s = str(cell).strip()
    if not s:
        return []
    # Reemplaza la coma decimal por punto cuando va seguida de 1-3 dígitos
    s = re.sub(r"(\d),(\d{1,3})\b", r"\1.\2", s)
    # Extrae todos los números (ya con punto decimal)
    nums = re.findall(r"-?\d+(?:\.\d+)?", s)
    out = []
    for n in nums:
        try:
            out.append(Decimal(n))
        except Exception:
            pass
    return out[:3]  # máximo 3 (uno por muestreo)


def parse_sabor(text):
    """Devuelve [(flavor_name, intensity_name, percentage)] (todos en MAYUS).

    Ejemplos:
      'CARACTERISTICO'         → [('CARACTERISTICO', None, None)]
      '40% TIERRA LEVE'        → [('TIERRA', 'LEVE', 40)]
      '70% TIERRA LEVE - 30% TIERRA MODERADA' → 2 entradas
    """
    if not text:
        return []
    s = str(text).strip()
    parts = re.split(r"\s*-\s*", s)
    out = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        m = re.match(
            r"^(\d+)\s*%\s*(.+?)(?:\s+(LEVE|MODERAD[OA]|FUERTE|NINGUNA))?$",
            p,
            re.IGNORECASE,
        )
        if m:
            pct = int(m.group(1))
            flavor = normalize(m.group(2))
            intensity = normalize(m.group(3)) if m.group(3) else None
            # Normaliza intensidad MODERADA/MODERADO → MODERADO (catálogo lo tiene como "Moderado")
            if intensity in ("MODERADA", "MODERADO"):
                intensity = "MODERADO"
            out.append((flavor, intensity, pct))
        else:
            out.append((normalize(p), None, None))
    return out


# ─────────────────────────────────────────────────────────────────────
# Find / find-or-create helpers (con caché en memoria)
# ─────────────────────────────────────────────────────────────────────

def find_plant(db, name):
    nm = normalize(name)
    if not nm:
        return None
    for p in db.execute(select(Plant)).scalars():
        if normalize(p.plant_name) == nm:
            return p
    return None


def foc_supplier(db, name, cache):
    nm = normalize(name)
    if not nm:
        return None
    if nm in cache:
        return cache[nm]
    for s in db.execute(select(Supplier)).scalars():
        if normalize(s.supplier_name) == nm:
            cache[nm] = s
            return s
    s = Supplier(supplier_name=nm)
    db.add(s)
    db.flush()
    cache[nm] = s
    cache.setdefault("__created__", []).append(("supplier", nm))
    return s


def foc_origin(db, name, cache):
    nm = normalize(name)
    if not nm:
        return None
    if nm in cache:
        return cache[nm]
    for o in db.execute(select(Origin)).scalars():
        if normalize(o.origin_name) == nm:
            cache[nm] = o
            return o
    o = Origin(origin_name=nm)
    db.add(o)
    db.flush()
    cache[nm] = o
    cache.setdefault("__created__", []).append(("origin", nm))
    return o


def foc_pond(db, code, supplier_id, origin_id, cache):
    if not code:
        return None
    nm = normalize(str(code))
    key = (nm, supplier_id, origin_id)
    if key in cache:
        return cache[key]
    q = select(Pond).where(Pond.pond_code == nm)
    if supplier_id:
        q = q.where(Pond.supplier_id == supplier_id)
    if origin_id:
        q = q.where(Pond.origin_id == origin_id)
    p = db.execute(q).scalar_one_or_none()
    if p:
        cache[key] = p
        return p
    p = Pond(pond_code=nm, supplier_id=supplier_id, origin_id=origin_id)
    db.add(p)
    db.flush()
    cache[key] = p
    cache.setdefault("__created__", []).append(("pond", nm))
    return p


def foc_logistics(db, name, cache):
    nm = normalize(name)
    if not nm:
        return None
    if nm in cache:
        return cache[nm]
    for x in db.execute(select(LogisticsCompany)).scalars():
        if normalize(x.company_name) == nm:
            cache[nm] = x
            return x
    x = LogisticsCompany(company_name=nm)
    db.add(x)
    db.flush()
    cache[nm] = x
    cache.setdefault("__created__", []).append(("logistics", nm))
    return x


def foc_treater(db, name, cache):
    nm = normalize(name)
    if not nm:
        return None
    if nm in cache:
        return cache[nm]
    for t in db.execute(select(Treater)).scalars():
        if normalize(t.full_name) == nm:
            cache[nm] = t
            return t
    t = Treater(full_name=nm, is_proveedor=(nm == "PROVEEDOR"))
    db.add(t)
    db.flush()
    cache[nm] = t
    cache.setdefault("__created__", []).append(("treater", nm))
    return t


def find_chemical(db, name):
    nm = normalize(name)
    if not nm:
        return None
    # Acepta variantes con paréntesis
    nm_no_spaces = nm.replace(" ", "")
    for c in db.execute(select(Chemical)).scalars():
        if normalize(c.chemical_name) == nm:
            return c
        if normalize(c.chemical_name).replace(" ", "") == nm_no_spaces:
            return c
    return None


def find_color(db, name):
    nm = normalize(name)
    if not nm:
        return None
    for c in db.execute(select(Color)).scalars():
        if normalize(c.color_name) == nm:
            return c
    return None


def find_flavor(db, name):
    nm = normalize(name)
    if not nm:
        return None
    for f in db.execute(select(Flavor)).scalars():
        if normalize(f.flavor_name) == nm:
            return f
    return None


def find_intensity(db, name):
    nm = normalize(name)
    if not nm:
        return None
    for i in db.execute(select(Intensity)).scalars():
        if normalize(i.intensity_name) == nm:
            return i
    return None


def load_defects(db):
    return {d.defect_code: d for d in db.execute(select(Defect)).scalars()}


# ─────────────────────────────────────────────────────────────────────
# Procesamiento por fila
# ─────────────────────────────────────────────────────────────────────

def process_row(db, row, defects_by_code, admin, sheet_name, row_idx,
                caches, stats, errors):
    sup_cache, ori_cache, pond_cache, log_cache, treater_cache = caches
    lote_cell = row[COLS["lote"]]
    if lote_cell is None or str(lote_cell).strip() == "":
        stats["empty_rows"] += 1
        return None
    lote_code = str(lote_cell).strip()

    fecha = parse_date_cell(row[COLS["fecha_recepcion"]])
    if not fecha:
        stats["skipped_invalid"] += 1
        errors.append(f"{sheet_name} f{row_idx}: lote {lote_code} sin fecha")
        return None

    lot_year = fecha.year

    existing = db.execute(
        select(Lot).where(Lot.lot_code == lote_code, Lot.lot_year == lot_year)
    ).scalar_one_or_none()
    if existing:
        stats["skipped_existing"] += 1
        return None

    planta_raw = normalize(row[COLS["planta"]])
    is_gerencia = planta_raw == "GERENCIA"
    plant = find_plant(db, row[COLS["planta"]])
    if not plant and is_gerencia:
        # "GERENCIA" no es planta, es categoría de lote. Default a CEA DURÁN.
        plant = find_plant(db, "CEA DURÁN")
    if not plant:
        stats["unmatched_plant"] += 1
        errors.append(f"{sheet_name} f{row_idx}: planta '{row[COLS['planta']]}' no en catálogo")
        return None
    # Si la fila era "GERENCIA", marcamos el lote como categoría gerencia
    lot_category_id = None
    if is_gerencia:
        cat = db.execute(
            select(LotCategory).where(LotCategory.category_code == "gerencia")
        ).scalar_one_or_none()
        if cat:
            lot_category_id = cat.lot_category_id

    supplier = foc_supplier(db, row[COLS["proveedor"]], sup_cache)
    origin = foc_origin(db, row[COLS["procedencia"]], ori_cache)
    pond = foc_pond(
        db, row[COLS["psc"]],
        supplier.supplier_id if supplier else None,
        origin.origin_id if origin else None,
        pond_cache,
    )
    logistics = foc_logistics(db, row[COLS["logistica"]], log_cache)
    chemical = find_chemical(db, row[COLS["metabisulfito"]])
    if row[COLS["metabisulfito"]] and not chemical:
        stats["unmatched_chemical"] += 1
    color = find_color(db, row[COLS["color"]])
    if row[COLS["color"]] and not color:
        stats["unmatched_color"] += 1

    sabor_entries = parse_sabor(row[COLS["sabor"]])
    treater_names = parse_treaters(row[COLS["tratador"]])
    treaters = [foc_treater(db, n, treater_cache) for n in treater_names]
    treaters = [t for t in treaters if t]

    tipo = normalize(row[COLS["tipo_producto"]])
    if tipo not in ("ENTERO", "COLA"):
        tipo = None

    obs = row[COLS["observacion"]]
    obs_str = str(obs).strip() if obs is not None and str(obs).strip() else None

    lot = Lot(
        lot_code=lote_code,
        lot_year=lot_year,
        plant_id=plant.plant_id,
        supplier_id=supplier.supplier_id if supplier else None,
        origin_id=origin.origin_id if origin else None,
        pond_id=pond.pond_id if pond else None,
        lot_category_id=lot_category_id,
        product_type=tipo,
        chemical_id=chemical.chemical_id if chemical else None,
        observations=obs_str,
        created_by=admin.user_id,
    )
    db.add(lot)
    db.flush()

    for t in treaters:
        db.execute(lot_treaters_table.insert().values(
            lot_id=lot.lot_id, treater_id=t.treater_id
        ))

    hora = parse_time_cell(row[COLS["hora_analisis"]])

    reception = Reception(
        plant_id=plant.plant_id,
        reception_date=fecha,
        arrival_time=hora,
        logistics_company_id=logistics.logistics_company_id if logistics else None,
        created_by=admin.user_id,
    )
    db.add(reception)
    db.flush()

    rl = ReceptionLot(
        reception_id=reception.reception_id,
        lot_id=lot.lot_id,
        sequence_in_reception=1,
        delivery_index=1,
        received_lbs=parse_decimal(row[COLS["lbs"]]),
    )
    db.add(rl)

    turno = normalize(row[COLS["turno"]])
    if turno not in ("T/D", "T/N"):
        turno = None

    pct_global = parse_decimal(row[COLS["pct_defectos"]])
    if pct_global is not None:
        # Excel da fracción (0.31). Nuestro campo está en percentage points (0-100).
        pct_global = pct_global * Decimal(100)

    # SO2: puede haber 1, 2 o 3 lecturas en la misma celda.
    so2_values = parse_so2_values(row[COLS["so2_global"]])
    so2_global_avg = (
        sum(so2_values) / len(so2_values) if so2_values else None
    )

    qa = QualityAnalysis(
        plant_id=plant.plant_id,
        analysis_date=fecha,
        analysis_time=hora,
        shift=turno,
        analyst_id=admin.user_id,
        gr_cc=parse_decimal(row[COLS["gr_cc"]]),
        c_kg=parse_decimal(row[COLS["c_kg"]]),
        gr_sc=parse_decimal(row[COLS["gr_sc"]]),
        c_kg2=parse_decimal(row[COLS["c_kg2"]]),
        so2_global=so2_global_avg,
        global_defect_percentage=pct_global,
        general_observations=obs_str,
        status="borrador",
        created_by=admin.user_id,
    )
    db.add(qa)
    db.flush()

    db.add(AnalysisLot(analysis_id=qa.analysis_id, lot_id=lot.lot_id))

    if color:
        db.add(AnalysisColor(
            analysis_id=qa.analysis_id, sample_state="cocido",
            color_id=color.color_id,
        ))

    for flavor_name, intensity_name, pct in sabor_entries:
        flavor = find_flavor(db, flavor_name)
        if not flavor:
            stats["unmatched_flavor"] += 1
            continue
        intensity = find_intensity(db, intensity_name) if intensity_name else None
        db.add(AnalysisFlavor(
            analysis_id=qa.analysis_id,
            sample_state="cocido",
            flavor_id=flavor.flavor_id,
            intensity_id=intensity.intensity_id if intensity else None,
            percentage=Decimal(pct) if pct is not None else None,
        ))

    # Muestreos: defectos van todos al 1er muestreo (Excel los trae agregados);
    # SO2 va distribuido según cuántas lecturas tenga la celda (1-3).
    has_defects = any(row[col] is not None for col in DEFECT_COL_MAP)
    needed_samplings = max(len(so2_values), 1 if has_defects else 0)

    samplings_by_idx = {}
    for i in range(1, needed_samplings + 1):
        s = AnalysisSampling(
            analysis_id=qa.analysis_id,
            sampling_index=i,
            so2_ppm=so2_values[i - 1] if i - 1 < len(so2_values) else None,
        )
        db.add(s)
        db.flush()
        samplings_by_idx[i] = s

    # Defectos siempre en el 1er muestreo (Excel los tiene agregados)
    if has_defects and 1 in samplings_by_idx:
        sampling1 = samplings_by_idx[1]
        for col_idx, defect_code in DEFECT_COL_MAP.items():
            v = row[col_idx]
            pct = parse_decimal(v)
            if pct is None or pct == 0:
                continue
            defect = defects_by_code.get(defect_code)
            if not defect:
                continue
            db.add(SamplingDefect(
                sampling_id=sampling1.sampling_id,
                defect_id=defect.defect_id,
                percentage=pct * Decimal(100),  # 0.05 → 5.00 percentage points
            ))

    return lot


# ─────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("file", help="Ruta al fichero Excel")
    parser.add_argument("--dry-run", action="store_true",
                        help="Simula sin commitear")
    parser.add_argument("--limit", type=int, default=None,
                        help="Solo procesar las N primeras filas (incluye saltadas)")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        print(f"[ERROR] No existe el fichero: {path}")
        sys.exit(1)

    print(f">> Leyendo {path}…")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"   Hojas: {', '.join(wb.sheetnames)}\n")

    db: Session = SessionLocal()
    try:
        admin = db.execute(select(User).order_by(User.user_id).limit(1)).scalar_one_or_none()
        if not admin:
            print("[ERROR] No hay usuarios. Ejecuta antes scripts/bootstrap_admin.py")
            sys.exit(1)
        print(f"   Importando como usuario: {admin.full_name} (#{admin.user_id})\n")

        defects_by_code = load_defects(db)
        if not defects_by_code:
            print("[ERROR] No hay defectos en el catálogo. Carga seeds.sql primero.")
            sys.exit(1)

        sup_cache = {}
        ori_cache = {}
        pond_cache = {}
        log_cache = {}
        treater_cache = {}
        caches = (sup_cache, ori_cache, pond_cache, log_cache, treater_cache)

        stats = {
            "imported": 0,
            "empty_rows": 0,
            "skipped_existing": 0,
            "skipped_invalid": 0,
            "unmatched_plant": 0,
            "unmatched_chemical": 0,
            "unmatched_color": 0,
            "unmatched_flavor": 0,
            "errors": 0,
        }
        errors = []
        rows_processed = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"=== Hoja: {sheet_name} (filas={ws.max_row}) ===")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i <= HEADER_ROW_IDX:
                    continue
                if args.limit is not None and rows_processed >= args.limit:
                    break
                try:
                    res = process_row(
                        db, row, defects_by_code, admin, sheet_name, i,
                        caches, stats, errors,
                    )
                    if res is not None:
                        stats["imported"] += 1
                        # Commit por fila: aísla fallos. Sin esto, un error
                        # rollback() destruye todo lo importado antes en la
                        # misma transacción.
                        if not args.dry_run:
                            db.commit()
                    else:
                        # No se importó (saltada). Liberar la sesión por si
                        # hubo INSERTs a catálogos (find-or-create).
                        if not args.dry_run:
                            db.commit()
                except Exception as e:
                    stats["errors"] += 1
                    errors.append(f"{sheet_name} f{i}: {type(e).__name__}: {str(e)[:200]}")
                    db.rollback()
                    # Limpiar cachés de objetos potencialmente expirados —
                    # se reconstruyen en la siguiente fila.
                    for c in caches:
                        c.clear()
                rows_processed += 1
                if rows_processed % 100 == 0:
                    print(f"  … procesadas {rows_processed} filas (importadas {stats['imported']})")
            if args.limit is not None and rows_processed >= args.limit:
                break

        if args.dry_run:
            print("\n[DRY-RUN] Rollback — nada se ha guardado.")
            db.rollback()
        else:
            print("\n[OK] Commits por fila aplicados.")

        print("\n=== RESUMEN ===")
        for k, v in stats.items():
            print(f"  {k:25} {v}")

        # Catálogos creados
        for cache, label in [
            (sup_cache, "suppliers"),
            (ori_cache, "origins"),
            (pond_cache, "ponds"),
            (log_cache, "logistics"),
            (treater_cache, "treaters"),
        ]:
            created = cache.get("__created__", [])
            if created:
                print(f"\n  {label} creados ({len(created)}):")
                for _, name in created[:30]:
                    print(f"    + {name}")
                if len(created) > 30:
                    print(f"    … y {len(created) - 30} más")

        if errors:
            print(f"\n[WARN]  {len(errors)} errores/saltos. Primeros 20:")
            for e in errors[:20]:
                print(f"    {e}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
